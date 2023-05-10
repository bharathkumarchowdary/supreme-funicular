import datetime
import re
import time

import yfinance as yf
import openpyxl
from openpyxl.reader.excel import load_workbook


def clear_content(sheet, start: str, end: str):
    for a in sheet[start:end]:  # you can set the range here
        for cell in a:
            cell.value = None  # set a value or null here


if __name__ == '__main__':
    '''
    
    '''

    date_q1 = '2022-06-30'
    date_q2 = '2022-09-30'
    date_q3 = '2022-12-31'
    date_q4 = '2023-03-30'

    # General options
    throttle_period = 5
    max_stocks = 200
    results_date = date_q4
    file_name = 'Nifty_100_Q4_FY23'
    # factor = 1e5  # Lakshalu
    factor = 1e7  # Kotlu
    heading = 'Q4: FY 23 (All Currency in Crores)'

    # Debug option
    # debug_stocks = True
    debug_stocks = False
    debug_stock_list = ['M&M']

    # Yfinance Library options
    # use_legacy_api = True
    use_legacy_api = False

    my_file = \
        '/Users/bharathkumargundala/Library/Mobile Documents/com~apple~CloudDocs/Stocks/Reports/' + file_name + '.xlsx'

    wb = load_workbook(filename=my_file)
    summary_sheet = wb['Summary']
    nifty_list_sheet = wb['ind_nifty100list']
    clear_content(summary_sheet, start='C4', end='C1000')
    clear_content(summary_sheet, start='E4', end='E1000')
    clear_content(summary_sheet, start='F4', end='F1000')
    clear_content(summary_sheet, start='H4', end='H1000')
    clear_content(summary_sheet, start='J4', end='J1000')
    index = 0
    offset = 4

    try:
        summary_sheet['D2'] = heading
        iterer = [row[2].value for row in
                  nifty_list_sheet.iter_rows(min_row=2)] if not debug_stocks else debug_stock_list
        for symbol in iterer:
            summary_index = str(index + offset)
            if index > max_stocks:
                break
            summary_sheet['C' + summary_index] = symbol
            stock = yf.Ticker(symbol + ".NS")
            if not use_legacy_api:
                qtly = stock.quarterly_incomestmt  # Recommended and latest use case
            else:
                qtly = stock.get_income_stmt(legacy=True, freq='quarterly')  # Legacy use case
            if results_date in qtly.columns:
                net_income = qtly[results_date]['Net Income Common Stockholders' if not use_legacy_api else 'NetIncome']
                total_revenue = qtly[results_date]['Total Revenue' if not use_legacy_api else 'TotalRevenue']
                summary_sheet['E' + summary_index] = net_income / factor
                summary_sheet['F' + summary_index] = total_revenue / factor
                if not use_legacy_api:
                    match = re.search(r'\d{4}-\d{2}-\d{2}', date_q2)
                    date_q2_p1 = datetime.datetime.strptime(match.group(), '%Y-%m-%d').date() + datetime.timedelta(
                        days=1)
                    price = stock.history(period='1d', interval='1d', start=date_q2, end=date_q2_p1, )['Close'][0]
                    diluted_eps = qtly[results_date]['Diluted EPS']
                    summary_sheet['H' + summary_index] = price / diluted_eps
                elif 'trailingPE' in stock.info:
                    summary_sheet['H' + summary_index] = stock.info['trailingPE'] / factor
                else:
                    summary_sheet['H' + summary_index] = 'NaN'
                summary_sheet['J' + summary_index] = stock.fast_info['market_cap'] / factor
                print('Quarterly Statement for ' + symbol + ' updated')
            else:
                summary_sheet['E' + summary_index] = 1e-3
                summary_sheet['F' + summary_index] = 1e-3
                summary_sheet['H' + summary_index] = 1e-3
                summary_sheet['J' + summary_index] = 1e-3
                print('Quarterly Statement for ' + symbol + ' not updated')
            index += 1
            time.sleep(throttle_period)

    finally:
        wb.save(my_file)
